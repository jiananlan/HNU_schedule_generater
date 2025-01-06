from openpyxl import load_workbook
import time
from datetime import date, datetime
import B9
import uuid

wb = load_workbook('jwxt-pkgl-axsdykb1.xlsx')
sheet = wb.active


class Mine:
    time_date: date
    time_begin: str
    time_end: str
    coure_name: str
    location: str
    teacher: str
    other: str

    def __init__(self, inpu: str, begin, end, day_):
        self.time_date = day_
        self.time_begin = begin
        self.time_end = end
        rs = inpu.split()
        self.coure_name = rs[0]
        self.teacher = rs[1]
        if len(rs) == 7:
            self.location = rs[4]
        else:
            self.location = ''
        self.other = f'{rs[-2]} {rs[-1]} {rs[2]}'

    def __str__(self):
        name_of_class, property_ = B9.pure_name(self.coure_name)
        x = str(datetime.now()).replace(" ", "").replace('-', '').replace(':', '').replace('.', '')[:-6]
        return f'''BEGIN:VEVENT
UID:{str(uuid.uuid4())}
DTSTAMP:{x[:8]}T{x[8:]}
DTSTART:{str(self.time_date).replace('-', '')}T{self.time_begin.replace(':', '')}00
DTEND:{str(self.time_date).replace('-', '')}T{self.time_end.replace(':', '')}00
SUMMARY:{name_of_class}
DESCRIPTION:{self.location}   {self.teacher} {property_} {self.other}
LOCATION:{B9.detect_location(self.location)}
END:VEVENT
'''


read_result = []

for row in sheet.iter_rows(min_row=4, min_col=2, max_col=8, max_row=8):
    for cell in row:
        result = str(cell.value).replace('\n', '').replace('  ', ' ')
        if result != 'None':
            read_result += result.split('-------------')

course_list = []

for row in read_result:
    r = row.split()[3]
    d, t = B9.main(r)
    for day in d:
        course_list.append(Mine(row, t[0], t[1], day))
    print(row)
del read_result


def write_main():
    with open('my1.ics', 'w', encoding='utf-8') as f:
        f.write('''BEGIN:VCALENDAR
VERSION:2.0
CALSCALE:GREGORIAN
''')
        for row_ in course_list:
            if not B9.is_holiday(row_.time_date):
                f.write(str(row_))
        f.write('END:VCALENDAR\n')


def write_shift():
    with open('my_shift1.ics', 'w', encoding='utf-8') as f:
        f.write('''BEGIN:VCALENDAR
VERSION:2.0
CALSCALE:GREGORIAN
''')
        for row_ in course_list:
            if row_.time_date in B9.DAY_SHIFT:
                row_.time_date = B9.DAY_SHIFT[row_.time_date]
                f.write(str(row_))
        f.write('END:VCALENDAR\n')


if __name__ == '__main__':
    write_main()
    write_shift()

print(time.process_time())
print(B9.error_count)
print(f'！第一天是{B9.FIRST_DAY}')
